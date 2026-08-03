#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
シニアジョブ(seniorjob.jp)の公開求人ページから、大阪府のシニア歓迎求人を
収集して RA(求人開拓)営業リスト(xlsx)に流し込む。

- 情報源: seniorjob.jp の公開一覧ページ /osaka/50s/ , /osaka/50s/page-N/
  (robots.txt 準拠。/apply/ /client/ /mypage/ 等や /*/page-0/ は取得しない)
- 各ページに埋め込まれた Nuxt データ(__NUXT_DATA__)を解析し、求人ごとの
  企業名・職種・給与・雇用形態・勤務地・特徴タグを構造化して取得する。
- status=public かつ企業名が公開されている(非公開求人でない)大阪府の求人のみ対象。
- 企業名で名寄せし1社1行。職種からカテゴリ(ホワイトカラー/配送系/工場系/その他)を判定。
- robots の ClaudeBot 向け Crawl-Delay:5 を尊重し、既定5秒間隔でアクセスする。

出力: scripts/out/RA営業リスト_大阪_シニアジョブ.xlsx

Usage:
  python3 scripts/seniorjob_prospect_collect.py \
      [--white 100] [--delivery 50] [--factory 50] [--other 50] \
      [--max-pages 60] [--delay 5.0] [--out <path>]
"""
import argparse
import json
import os
import re
import sys
import time
import html

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

BASE = "https://seniorjob.jp"
LIST_PATH = "/osaka/50s"  # 大阪府 × 50歳以上
UA = "SeniorJob-RA-list/1.0 (ClaudeBot compatible; internal recruitment tool)"

# カテゴリ判定(配送→工場→ホワイトカラーの順で職種名/タイトルを照合、外れれば その他)
DELIVERY_KW = ["ドライバー", "配送", "運搬", "配達", "トラック", "軽貨物",
               "ルート配", "宅配", "送迎", "運転"]
FACTORY_KW = ["製造", "検品", "組立", "組み立て", "加工", "ライン", "梱包",
              "フォークリフト", "倉庫", "工場", "溶接", "機械オペ", "品質管理",
              "塗装", "プレス", "旋盤", "工員", "ものづくり", "包装"]
WHITE_KW = ["事務", "経理", "総務", "営業", "人事", "受付", "データ入力",
            "会計", "税理", "経営", "企画", "コールセンター", "テレア",
            "オペレーター", "CAD", "設計", "施工管理", "貿易", "秘書", "広報",
            "マーケ", "エンジニア", "プログラ", "行政書士", "社労士", "宅建",
            "士業", "管理業務", "コンサル", "IT"]

SENIOR_TAGS = ("シニア", "50代", "60代", "70代", "年齢不問", "中高年", "ミドル")

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
LABEL_FILL = PatternFill("solid", fgColor="D9E1F2")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = [
    ("No.", 6), ("業種カテゴリ", 14), ("企業名", 28), ("募集職種", 26),
    ("都道府県", 10), ("市区町村", 16), ("掲載媒体", 12), ("シニア歓迎", 10),
    ("電話番号", 16), ("問い合わせURL/メール", 26), ("掲載件数", 9),
    ("優先度", 8), ("ステータス", 12), ("送客候補メモ", 30), ("最終接触日", 12),
    ("備考", 24),
]
DV = {
    "業種カテゴリ": "ホワイトカラー,配送系,工場系,その他",
    "掲載媒体": "シニアジョブ,求人ボックス,Indeed,スタンバイ,Engage,その他",
    "シニア歓迎": "○,△,×,不明",
    "優先度": "A,B,C",
    "ステータス": "未着手,架電済,アポ,送客,見送り,失注",
}

CATEGORIES = ["ホワイトカラー", "配送系", "工場系", "その他"]


def fetch(session, page):
    if page <= 1:
        url = f"{BASE}{LIST_PATH}/"
    else:
        url = f"{BASE}{LIST_PATH}/page-{page}/"
    r = session.get(url, timeout=30)
    if r.status_code == 404:
        return None
    r.raise_for_status()
    return r.text


def extract_jobs(page_html):
    """__NUXT_DATA__(devalue)を解析して求人オブジェクト配列を返す。"""
    m = re.search(r'id="__NUXT_DATA__"[^>]*>(.*?)</script>', page_html, re.S)
    if not m:
        return []
    try:
        arr = json.loads(html.unescape(m.group(1)))
    except Exception:
        return []

    sys.setrecursionlimit(100000)

    def R(i, seen):
        if not isinstance(i, int) or i < 0 or i >= len(arr):
            return i
        if i in seen:
            return None
        v = arr[i]
        if isinstance(v, dict):
            return {k: R(vi, seen | {i}) for k, vi in v.items()}
        if isinstance(v, list):
            return [R(x, seen | {i}) for x in v]
        return v

    jobs, ids = [], set()
    for i, v in enumerate(arr):
        if isinstance(v, dict) and "title" in v and "company" in v and "workLocationsText" in v:
            j = R(i, set())
            jid = j.get("id")
            if jid in ids:
                continue
            ids.add(jid)
            jobs.append(j)
    return jobs


def city_of(location):
    if "大阪" not in (location or ""):
        return None
    m = re.search(r"大阪府\s*(大阪市[^\s0-9/]*区|[^\s0-9/]+市|[^\s0-9/]+区|[^\s0-9/]+町|[^\s0-9/]+郡[^\s0-9/]*)", location)
    return m.group(1) if m else "大阪府内"


def categorize(occupation, title):
    text = f"{occupation} {title}"
    for kw in DELIVERY_KW:
        if kw in text:
            return "配送系"
    for kw in FACTORY_KW:
        if kw in text:
            return "工場系"
    for kw in WHITE_KW:
        if kw in text:
            return "ホワイトカラー"
    return "その他"


def senior_flag(tags):
    joined = " ".join(tags or [])
    if any(t in joined for t in SENIOR_TAGS):
        return "○"
    if "ブランクOK" in joined:
        return "△"
    return "不明"


def collect(session, targets, max_pages, delay):
    buckets = {c: [] for c in CATEGORIES}
    seen_company = set()
    page = 1
    retries = 0
    MAX_RETRIES = 3
    while page <= max_pages:
        if all(len(buckets[c]) >= targets[c] for c in CATEGORIES):
            break
        try:
            page_html = fetch(session, page)
        except Exception as e:  # noqa
            if "429" in str(e) and retries < MAX_RETRIES:
                retries += 1
                wait = min(delay * 2 ** retries, 60)
                print(f"  ! page {page}: 429 レート制限、{wait:.0f}s 待機して再試行 ({retries}/{MAX_RETRIES})", file=sys.stderr)
                time.sleep(wait)
                continue  # 同じページを再試行
            print(f"  ! page {page}: {e}(中止)", file=sys.stderr)
            break
        retries = 0
        if page_html is None:
            break
        jobs = extract_jobs(page_html)
        if not jobs:
            print(f"  page {page}: 求人0件、終了")
            break
        added = 0
        for j in jobs:
            comp = (j.get("company") or {}).get("name")
            loc = j.get("workLocationsText") or ""
            if not comp or j.get("status") != "public":
                continue
            city = city_of(loc)
            if city is None:
                continue
            key = re.sub(r"\s+", "", comp)
            if key in seen_company:
                continue
            occ = (j.get("occupation") or {}).get("name") or ""
            title = j.get("title") or ""
            cat = categorize(occ, title)
            if len(buckets[cat]) >= targets[cat]:
                continue  # この分類は充足済み(他分類の枠で拾える場合のみ採用)
            seen_company.add(key)
            tags = j.get("featureTags") or []
            memo_bits = [b for b in [j.get("salaryText"),
                                     (j.get("employmentType") or {}).get("name")] if b]
            if tags:
                memo_bits.append("／".join(tags[:6]))
            buckets[cat].append({
                "company": comp, "title": occ or title, "city": city,
                "senior": senior_flag(tags), "memo": " / ".join(memo_bits),
            })
            added += 1
        counts = " ".join(f"{c[:2]}{len(buckets[c])}" for c in CATEGORIES)
        print(f"  page {page}: +{added} ({counts})")
        time.sleep(delay)
    return buckets


def build_workbook(buckets, targets):
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("リスト")
    for i, (h, w) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = w
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    idx = {h: i + 1 for i, (h, _) in enumerate(COLUMNS)}

    r, no = 2, 1
    for cat in CATEGORIES:
        for row in buckets.get(cat, []):
            vals = {
                "No.": no, "業種カテゴリ": cat, "企業名": row["company"],
                "募集職種": row["title"], "都道府県": "大阪府", "市区町村": row["city"],
                "掲載媒体": "シニアジョブ", "シニア歓迎": row["senior"],
                "ステータス": "未着手", "備考": row["memo"],
            }
            for h, _ in COLUMNS:
                cell = ws.cell(row=r, column=idx[h], value=vals.get(h, ""))
                cell.font = Font(name=FONT, size=10)
                cell.border = BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if h in ("No.", "業種カテゴリ", "都道府県"):
                    cell.fill = LABEL_FILL
            no += 1
            r += 1
    last = r - 1
    for h, opts in DV.items():
        dv = DataValidation(type="list", formula1=f'"{opts}"', allow_blank=True)
        ws.add_data_validation(dv)
        col = get_column_letter(idx[h])
        if last >= 2:
            dv.add(f"{col}2:{col}{last}")

    s = wb.create_sheet("説明", 0)
    s.sheet_view.showGridLines = False
    s.column_dimensions["A"].width = 20
    s.column_dimensions["B"].width = 82
    t = s.cell(row=1, column=1, value="RA営業リスト(大阪) — シニアジョブ収集版")
    t.font = Font(name=FONT, bold=True, size=15, color="1F4E78")
    lines = [
        ("情報源", "シニアジョブ(seniorjob.jp) 公開一覧 /osaka/50s/。50歳以上歓迎・企業名公開の求人のみ収集。"),
        ("エリア", "大阪府(市区町村は勤務地表記から抽出)。"),
        ("収集件数", " / ".join(f"{c}:{len(buckets.get(c, []))}/{targets[c]}" for c in CATEGORIES)),
        ("dedup", "企業名で名寄せし1社1行。"),
        ("カテゴリ", "募集職種から判定(配送→工場→ホワイトカラーの優先順、外れれば その他)。"),
        ("シニア歓迎", "求人の特徴タグから判定(○=シニア/50〜70代/年齢不問、△=ブランクOK)。"),
        ("robots", "公開一覧ページのみ取得。会員/応募/企業管理ページや page-0 は取得せず、Crawl-Delay:5 を遵守。"),
        ("連絡先", "電話・問い合わせURLは未取得(空欄)。各社公式サイト等から転記してください。"),
        ("留意", "掲載情報は変動します。架電前に最新掲載で企業名/連絡先を再確認してください。"),
    ]
    rr = 3
    for k, v in lines:
        s.cell(row=rr, column=1, value=k).font = Font(name=FONT, bold=True, size=10)
        c = s.cell(row=rr, column=2, value=v)
        c.font = Font(name=FONT, size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        rr += 1
    wb.active = 0
    return wb


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--white", type=int, default=100)
    ap.add_argument("--delivery", type=int, default=50)
    ap.add_argument("--factory", type=int, default=50)
    ap.add_argument("--other", type=int, default=50)
    ap.add_argument("--max-pages", type=int, default=60)
    ap.add_argument("--delay", type=float, default=5.0)
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    targets = {"ホワイトカラー": args.white, "配送系": args.delivery,
               "工場系": args.factory, "その他": args.other}

    session = requests.Session()
    session.headers.update({"User-Agent": UA, "Accept-Language": "ja,en;q=0.8"})
    ca = "/root/.ccr/ca-bundle.crt"
    if os.path.exists(ca):
        session.verify = ca

    print(f"== シニアジョブ 大阪×50代以上 収集 (delay {args.delay}s) ==")
    buckets = collect(session, targets, args.max_pages, args.delay)

    out = args.out or os.path.join(os.path.dirname(__file__), "out",
                                   "RA営業リスト_大阪_シニアジョブ.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb = build_workbook(buckets, targets)
    wb.save(out)
    total = sum(len(v) for v in buckets.values())
    print(f"\nsaved: {out}  (計{total}社)")
    for c in CATEGORIES:
        print(f"  {c}: {len(buckets[c])}")


if __name__ == "__main__":
    main()
