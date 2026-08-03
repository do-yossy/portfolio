#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
収集済みの RA営業リスト(各媒体xlsx)を企業名で名寄せして1ファイルに統合する。

入力(存在するものだけ使用):
  scripts/out/RA営業リスト_大阪_シニア求人ナビ.xlsx   (企業名+電話+シニア)
  scripts/out/RA営業リスト_大阪_求人ボックス.xlsx     (企業名・ボリューム)
出力:
  scripts/out/RA営業リスト_大阪_統合.xlsx

- 企業名で名寄せ(1社1行)。複数媒体で重複した場合は媒体名を併記し、電話番号など
  情報量の多い方を優先採用。
- カテゴリ(ホワイトカラー/配送系/工場系/その他)ごとに並べる。
- 「情報源サマリ」タブに各サイトの収集可否を記載。
"""
import os
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(__file__)
OUT_DIR = os.path.join(HERE, "out")
# 優先度の高い(情報量の多い)媒体を先に置く
SOURCES = [
    ("シニア求人ナビ", "RA営業リスト_大阪_シニア求人ナビ.xlsx"),
    ("求人ボックス", "RA営業リスト_大阪_求人ボックス.xlsx"),
]

COLUMNS = [
    ("No.", 6), ("業種カテゴリ", 14), ("企業名", 28), ("募集職種", 30),
    ("都道府県", 10), ("市区町村", 16), ("掲載媒体", 18), ("シニア歓迎", 10),
    ("電話番号", 16), ("問い合わせURL/メール", 24), ("掲載件数", 9),
    ("優先度", 8), ("ステータス", 12), ("送客候補メモ", 30), ("最終接触日", 12),
    ("備考", 26),
]
CATEGORIES = ["ホワイトカラー", "配送系", "工場系", "その他"]

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
LABEL_FILL = PatternFill("solid", fgColor="D9E1F2")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
DV = {
    "業種カテゴリ": "ホワイトカラー,配送系,工場系,その他",
    "掲載媒体": "シニア求人ナビ,求人ボックス,シニアジョブ,Indeed,その他",
    "シニア歓迎": "○,△,×,不明",
    "優先度": "A,B,C",
    "ステータス": "未着手,架電済,アポ,送客,見送り,失注",
}
SENIOR_RANK = {"○": 3, "△": 2, "不明": 1, "×": 0, None: 0, "": 0}


def norm(name):
    return re.sub(r"\s|株式会社|有限会社|合同会社|（株）|\(株\)", "", name or "")


def read_rows(path):
    wb = load_workbook(path)
    ws = wb["リスト"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    hidx = {h: i + 1 for i, h in enumerate(headers)}
    rows = []
    for r in range(2, ws.max_row + 1):
        comp = ws.cell(r, hidx.get("企業名", 3)).value
        if not comp:
            continue
        rows.append({h: ws.cell(r, hidx[h]).value for h in headers if h in hidx})
    return rows


def merge():
    merged = {}  # norm(company) -> row dict
    counts = {}
    for media, fname in SOURCES:
        path = os.path.join(OUT_DIR, fname)
        if not os.path.exists(path):
            counts[media] = 0
            continue
        rows = read_rows(path)
        counts[media] = len(rows)
        for row in rows:
            key = norm(row.get("企業名"))
            if not key:
                continue
            row = dict(row)
            row["掲載媒体"] = media
            if key not in merged:
                merged[key] = row
                continue
            # 既存とマージ: 電話・シニア判定など情報量の多い方を採用
            cur = merged[key]
            if not cur.get("電話番号") and row.get("電話番号"):
                cur["電話番号"] = row["電話番号"]
            if SENIOR_RANK.get(row.get("シニア歓迎"), 0) > SENIOR_RANK.get(cur.get("シニア歓迎"), 0):
                cur["シニア歓迎"] = row["シニア歓迎"]
            if (cur.get("業種カテゴリ") == "その他") and row.get("業種カテゴリ") not in (None, "その他"):
                cur["業種カテゴリ"] = row["業種カテゴリ"]
            media_set = set(str(cur.get("掲載媒体", "")).split("/")) | {media}
            cur["掲載媒体"] = "/".join(sorted(m for m in media_set if m))
            # メモ/備考は空なら補完
            for f in ("送客候補メモ", "備考", "募集職種"):
                if not cur.get(f) and row.get(f):
                    cur[f] = row[f]
    return merged, counts


def build(merged, counts):
    wb = Workbook()
    wb.remove(wb.active)

    # サマリタブ
    s = wb.create_sheet("情報源サマリ")
    s.sheet_view.showGridLines = False
    s.column_dimensions["A"].width = 18
    s.column_dimensions["B"].width = 12
    s.column_dimensions["C"].width = 74
    t = s.cell(row=1, column=1, value="RA営業リスト(大阪) — 統合版 / 情報源サマリ")
    t.font = Font(name=FONT, bold=True, size=15, color="1F4E78")
    hdr = ["情報源", "収集可否", "備考"]
    for i, h in enumerate(hdr, 1):
        c = s.cell(row=3, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER
    src_notes = [
        ("シニア求人ナビ", "○ 収集済", f"seniorjob-navi.com。企業名+電話+シニア歓迎。大阪×正社員 {counts.get('シニア求人ナビ', 0)}社(在庫上限)。"),
        ("求人ボックス", "○ 収集済", f"kyujinbox。企業名あり・ボリューム大。{counts.get('求人ボックス', 0)}社。"),
        ("シニアジョブ(seniorjob.jp)", "△ 制限", "企業名あり。ただしアクセスがレート制限(429)され当環境では大量取得不可。スクリプトは同梱。"),
        ("リクナビNEXT", "× 不可", "robots.txt で当該検索(?prf= ?emp= ?l1=)を全面Disallow。収集対象外。"),
        ("ミドルの転職", "× 不可", "robots.txt(User-agent:*)で /wish-search/ /company/ /job-search/ をDisallow。加えてbot遮断(403)。"),
    ]
    rr = 4
    for name, ok, note in src_notes:
        s.cell(row=rr, column=1, value=name).font = Font(name=FONT, size=10, bold=True)
        cok = s.cell(row=rr, column=2, value=ok)
        cok.font = Font(name=FONT, size=10)
        cok.alignment = Alignment(horizontal="center")
        cn = s.cell(row=rr, column=3, value=note)
        cn.font = Font(name=FONT, size=10)
        cn.alignment = Alignment(wrap_text=True, vertical="top")
        for cc in range(1, 4):
            s.cell(row=rr, column=cc).border = BORDER
        rr += 1
    rr += 1
    total = len(merged)
    tel = sum(1 for v in merged.values() if v.get("電話番号"))
    for line in [
        f"統合後の企業数(名寄せ後): {total}社",
        f"うち直通電話あり: {tel}社",
        "留意: 掲載情報は変動します。架電前に各社公式・最新掲載で企業名/連絡先を再確認してください。",
        "電話・問い合わせURLが空欄の社は、各社公式サイト等から転記してください。",
    ]:
        c = s.cell(row=rr, column=1, value=line)
        c.font = Font(name=FONT, size=10)
        s.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=3)
        rr += 1

    # リストタブ
    ws = wb.create_sheet("リスト")
    for i, (h, w) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = w
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    idx = {h: i + 1 for i, (h, _) in enumerate(COLUMNS)}

    def sortkey(v):
        cat = v.get("業種カテゴリ") or "その他"
        ci = CATEGORIES.index(cat) if cat in CATEGORIES else 9
        has_tel = 0 if v.get("電話番号") else 1
        return (ci, has_tel, v.get("企業名") or "")

    r, no = 2, 1
    for v in sorted(merged.values(), key=sortkey):
        vals = dict(v)
        vals["No."] = no
        if not vals.get("ステータス"):
            vals["ステータス"] = "未着手"
        if not vals.get("都道府県"):
            vals["都道府県"] = "大阪府"
        for h, _ in COLUMNS:
            cell = ws.cell(row=r, column=idx[h], value=vals.get(h, ""))
            cell.font = Font(name=FONT, size=10)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if h in ("No.", "業種カテゴリ", "都道府県"):
                cell.fill = LABEL_FILL
        no += 1
        r += 1
    last = r - 1
    for h, opts in DV.items():
        dv = DataValidation(type="list", formula1=f'"{opts}"', allow_blank=True)
        ws.add_data_validation(dv)
        col = get_column_letter(idx[h])
        if last >= 2:
            dv.add(f"{col}2:{col}{last}")
    return wb


def main():
    merged, counts = merge()
    wb = build(merged, counts)
    os.makedirs(OUT_DIR, exist_ok=True)
    out = os.path.join(OUT_DIR, "RA営業リスト_大阪_統合.xlsx")
    wb.save(out)
    print(f"saved: {out}")
    print(f"  統合後 {len(merged)}社 / 直通電話 {sum(1 for v in merged.values() if v.get('電話番号'))}社")
    for m, n in counts.items():
        print(f"  入力 {m}: {n}社")


if __name__ == "__main__":
    main()
