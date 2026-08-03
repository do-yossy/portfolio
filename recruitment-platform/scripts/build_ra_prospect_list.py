#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
RA(求人開拓)営業リストの器(テンプレート)を生成する。
大阪府・カテゴリ別200行(ホワイトカラー100 / 配送系50 / 工場系50)。
企業名・連絡先は空欄。実在ソース(Indeed Publisher API 等)から流し込む前提。
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
FILLIN_FILL = PatternFill("solid", fgColor="FFF2CC")   # 黄=手入力/流し込み対象
LABEL_FILL = PatternFill("solid", fgColor="D9E1F2")     # 事前ラベル(編集不要)
SAMPLE_FILL = PatternFill("solid", fgColor="E2EFDA")    # サンプル行
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# --- カテゴリ配分(大阪府) ---
CATEGORIES = [("ホワイトカラー", 100), ("配送系", 50), ("工場系", 50)]

# --- 列定義: (見出し, 幅, 種別)  種別: label=事前ラベル / fill=手入力 ---
COLUMNS = [
    ("No.", 6, "label"),
    ("業種カテゴリ", 14, "label"),
    ("企業名", 26, "fill"),
    ("募集職種", 22, "fill"),
    ("都道府県", 10, "label"),
    ("市区町村", 14, "fill"),
    ("掲載媒体", 14, "fill"),
    ("シニア歓迎", 10, "fill"),
    ("電話番号", 16, "fill"),
    ("問い合わせURL/メール", 28, "fill"),
    ("掲載件数", 9, "fill"),
    ("優先度", 8, "fill"),
    ("ステータス", 12, "fill"),
    ("送客候補メモ", 24, "fill"),
    ("最終接触日", 12, "fill"),
    ("備考", 20, "fill"),
]

# ドロップダウン候補
DV = {
    "業種カテゴリ": "ホワイトカラー,配送系,工場系",
    "掲載媒体": "シニアジョブ,求人ボックス,Indeed,スタンバイ,Engage,その他",
    "シニア歓迎": "○,△,×,不明",
    "優先度": "A,B,C",
    "ステータス": "未着手,架電済,アポ,送客,見送り,失注",
}

# カテゴリ別: 想定職種 と 収集キーワード例
CATEGORY_GUIDE = {
    "ホワイトカラー": {
        "職種例": "一般事務 / 経理 / 総務 / 営業事務 / 受付 / データ入力 / コールセンター / カスタマーサポート / 管理部門",
        "検索KW例": "大阪 事務 シニア歓迎 / 大阪 経理 60歳以上 / 大阪 データ入力 未経験",
    },
    "配送系": {
        "職種例": "軽貨物ドライバー / ルート配送 / 配送助手 / 宅配 / 倉庫内ピッキング / 仕分け / センター内作業",
        "検索KW例": "大阪 配送 ドライバー シニア / 大阪 ルート配送 中高年 / 大阪 仕分け 短時間",
    },
    "工場系": {
        "職種例": "製造ライン / 検品 / 組立 / 加工 / 梱包 / 設備保全 / フォークリフト / 品質管理",
        "検索KW例": "大阪 工場 製造 シニア歓迎 / 大阪 検品 軽作業 / 大阪 フォークリフト 60代",
    },
}


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def build_list_sheet(wb):
    ws = wb.create_sheet("リスト")
    headers = [c[0] for c in COLUMNS]
    for i, (h, w, _) in enumerate(COLUMNS, start=1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = w
    style_header(ws, len(COLUMNS))

    col_idx = {h: i + 1 for i, (h, _, _) in enumerate(COLUMNS)}

    # --- サンプル行(1行だけ・形式見本) ---
    sample = {
        "No.": "例",
        "業種カテゴリ": "ホワイトカラー",
        "企業名": "（例）株式会社サンプル商事",
        "募集職種": "一般事務",
        "都道府県": "大阪府",
        "市区町村": "大阪市中央区",
        "掲載媒体": "シニアジョブ",
        "シニア歓迎": "○",
        "電話番号": "06-0000-0000",
        "問い合わせURL/メール": "https://example.co.jp/contact",
        "掲載件数": 3,
        "優先度": "A",
        "ステータス": "未着手",
        "送客候補メモ": "60代・事務経験者を想定",
        "最終接触日": "2026/08/03",
        "備考": "※これは形式見本の行です。実データ入力時は削除可",
    }
    for h, v in sample.items():
        cell = ws.cell(row=2, column=col_idx[h], value=v)
        cell.font = Font(name=FONT, italic=True, size=10, color="548235")
        cell.fill = SAMPLE_FILL
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    # --- 本体200行(大阪府・カテゴリ事前ラベル) ---
    start = 3
    row = start
    no = 1
    label_cols = {c[0] for c in COLUMNS if c[2] == "label"}
    for cat, count in CATEGORIES:
        for _ in range(count):
            for h, _, kind in COLUMNS:
                cell = ws.cell(row=row, column=col_idx[h])
                cell.border = BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if h == "No.":
                    cell.value = no
                    cell.fill = LABEL_FILL
                    cell.font = Font(name=FONT, size=10)
                elif h == "業種カテゴリ":
                    cell.value = cat
                    cell.fill = LABEL_FILL
                    cell.font = Font(name=FONT, size=10)
                elif h == "都道府県":
                    cell.value = "大阪府"
                    cell.fill = LABEL_FILL
                    cell.font = Font(name=FONT, size=10)
                else:
                    cell.fill = FILLIN_FILL
                    cell.font = Font(name=FONT, size=10)
            no += 1
            row += 1
    last = row - 1

    # --- データ検証(ドロップダウン)を本体範囲に適用 ---
    for h, opts in DV.items():
        dv = DataValidation(type="list", formula1=f'"{opts}"', allow_blank=True)
        dv.error = "リストから選択してください"
        dv.errorTitle = "入力値エラー"
        ws.add_data_validation(dv)
        col = get_column_letter(col_idx[h])
        dv.add(f"{col}{start}:{col}{last}")

    return ws, last


def build_guide_sheet(wb):
    ws = wb.create_sheet("説明", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 90

    def h(text, r):
        c = ws.cell(row=r, column=1, value=text)
        c.font = Font(name=FONT, bold=True, size=13, color="1F4E78")
        return r + 1

    def kv(k, v, r):
        ws.cell(row=r, column=1, value=k).font = Font(name=FONT, bold=True, size=10)
        cell = ws.cell(row=r, column=2, value=v)
        cell.font = Font(name=FONT, size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        return r + 1

    r = 1
    t = ws.cell(row=r, column=1, value="RA(求人開拓)営業リスト — 大阪府")
    t.font = Font(name=FONT, bold=True, size=16, color="1F4E78")
    r += 2
    r = kv("目的", "シニア層を送客できる大阪府の採用中企業を集め、RA(法人)営業の架電・アプローチに使う。", r)
    r = kv("対象件数", "合計200件 / ホワイトカラー100・配送系50・工場系50(『リスト』タブに事前ラベル済)。", r)
    r = kv("エリア", "大阪府(市区町村は各行で入力)。", r)
    r += 1

    r = h("【重要】企業データの入れ方", r)
    r = kv("原則", "企業名・電話・連絡先などの実データは必ず実在ソースから入れること。創作・推測で埋めない。", r)
    r = kv("推奨ソース①", "Indeed Publisher API(lib/indeed-api.js searchJobs)。q=職種, l=大阪 で company/title/location/url を取得。INDEED_PUBLISHER_ID が必要。", r)
    r = kv("推奨ソース②", "シニアジョブ等のシニア特化媒体・求人ボックス・スタンバイの掲載企業(各媒体の利用規約・robotsを確認のうえ)。", r)
    r = kv("連絡先の扱い", "電話・問い合わせ先は各社の公式サイト/掲載情報から確認して転記。個人情報の取り扱いに注意。", r)
    r += 1

    r = h("入力ルール(色分け)", r)
    for label, fill, desc in [
        ("黄色セル", FILLIN_FILL, "手入力/流し込み対象(企業名・職種・連絡先・媒体・ステータス等)"),
        ("水色セル", LABEL_FILL, "事前ラベル(No./業種カテゴリ/都道府県=大阪府)。基本は編集不要"),
        ("緑色セル", SAMPLE_FILL, "形式見本の行(2行目)。実データ入力時は削除してよい"),
    ]:
        cell = ws.cell(row=r, column=1, value=label)
        cell.fill = fill
        cell.font = Font(name=FONT, bold=True, size=10)
        cell.border = BORDER
        d = ws.cell(row=r, column=2, value=desc)
        d.font = Font(name=FONT, size=10)
        d.alignment = Alignment(wrap_text=True, vertical="center")
        r += 1
    r += 1

    r = h("ドロップダウン付き列", r)
    r = kv("業種カテゴリ", "ホワイトカラー / 配送系 / 工場系", r)
    r = kv("掲載媒体", "シニアジョブ / 求人ボックス / Indeed / スタンバイ / Engage / その他", r)
    r = kv("シニア歓迎", "○ / △ / × / 不明", r)
    r = kv("優先度", "A / B / C(掲載件数が多い=採用ニーズ高=A の目安)", r)
    r = kv("ステータス", "未着手 / 架電済 / アポ / 送客 / 見送り / 失注", r)
    r += 1

    r = h("カテゴリ別 想定職種 & 収集キーワード例", r)
    for cat in ("ホワイトカラー", "配送系", "工場系"):
        g = CATEGORY_GUIDE[cat]
        cell = ws.cell(row=r, column=1, value=cat)
        cell.font = Font(name=FONT, bold=True, size=11, color="1F4E78")
        r += 1
        r = kv("  想定職種", g["職種例"], r)
        r = kv("  検索KW例", g["検索KW例"], r)
        r += 1

    ws.cell(row=r, column=1, value="※本ファイルは『器(テンプレート)』です。実データはリポジトリの収集フロー(Indeed API 等)または手作業で流し込んでください。").font = Font(name=FONT, italic=True, size=9, color="808080")
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)
    build_list_sheet(wb)
    build_guide_sheet(wb)
    wb.active = 0

    out_dir = os.path.join(os.path.dirname(__file__), "out")
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, "RA営業リスト_大阪_テンプレート.xlsx")
    wb.save(path)
    print("saved:", path)


if __name__ == "__main__":
    main()
