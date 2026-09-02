#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FROM40（from-40.jp／40代・50代の転職）の公開求人から、大阪府×正社員の企業を
収集して RA(求人開拓)営業リスト(xlsx)に流し込む。

- 情報源: https://www.from-40.jp/recruitments?s[pref_id]=27&s[employment_ids][]=1
  （pref_id=27=大阪府 / employment=1=正社員）。robots.txt 制限なし。
- 一覧から 企業名・会社ページID・募集職種 を抽出し、企業名で名寄せ（1社1行）。
- 各社の会社ページ /companies/{id} から 所在地・公式サイトURL・事業内容・代表者・
  従業員数 を取得（電話番号は会社ページには無いため公式サイトURLを連絡導線として記載）。
- 職種/事業内容からカテゴリ（ホワイトカラー/配送系/工場系/その他）を判定。

出力: scripts/out/RA営業リスト_大阪_FROM40.xlsx

Usage:
  python3 scripts/from40_prospect_collect.py [--max-pages 25] [--max-companies 200] [--delay 1.2] [--out <path>]
"""
import argparse
import os
import re
import sys
import time
import html

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

BASE = "https://www.from-40.jp"
LIST_PATH = "/recruitments?s%5Bpref_id%5D=27&s%5Bemployment_ids%5D%5B%5D=1"
UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/120.0 Safari/537.36")

CARD_RE = re.compile(
    r'href="/companies/(\d+)"[^>]*>([^<]+)</a>\s*<h4>\s*<a[^>]*href="/recruitments/(\d+)"[^>]*>([^<]+)</a>',
    re.S)

DELIVERY_KW = ["ドライバー", "配送", "運搬", "配達", "トラック", "軽貨物", "ルート配",
               "宅配", "送迎", "運転手", "乗務員", "タクシー", "運送", "運輸"]
FACTORY_KW = ["製造", "検品", "組立", "加工", "ライン", "梱包", "フォークリフト", "倉庫",
              "工場", "溶接", "機械オペ", "品質管理", "塗装", "プレス", "工員", "整備", "生産"]
WHITE_KW = ["事務", "経理", "総務", "営業", "人事", "受付", "データ入力", "会計", "税理",
            "経営", "企画", "コールセンター", "オペレーター", "CAD", "設計", "施工管理",
            "貿易", "秘書", "広報", "マーケ", "エンジニア", "プログラ", "SE", "IT",
            "コンサル", "管理", "販売", "接客"]
SENIOR_KW = ("シニア", "40代", "50代", "60代", "年齢不問", "中高年", "ミドル")

CATEGORIES = ["ホワイトカラー", "配送系", "工場系", "その他"]

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
LABEL_FILL = PatternFill("solid", fgColor="D9E1F2")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = [
    ("No.", 6), ("業種カテゴリ", 14), ("企業名", 30), ("募集職種", 30),
    ("都道府県", 10), ("市区町村", 16), ("掲載媒体", 12), ("シニア歓迎", 10),
    ("電話番号", 14), ("問い合わせURL/メール", 30), ("掲載件数", 9),
    ("優先度", 8), ("ステータス", 12), ("送客候補メモ", 30), ("最終接触日", 12),
    ("備考", 34),
]
DV = {
    "業種カテゴリ": "ホワイトカラー,配送系,工場系,その他",
    "掲載媒体": "FROM40,シニア求人ナビ,求人ボックス,Indeed,その他",
    "シニア歓迎": "○,△,×,不明",
    "優先度": "A,B,C",
    "ステータス": "未着手,架電済,アポ,送客,見送り,失注",
}


def fetch(session, url):
    r = session.get(url, timeout=30)
    if r.status_code in (404, 410):
        return None
    r.raise_for_status()
    return r.text


def clean(s):
    return re.sub(r"\s+", " ", html.unescape(re.sub(r"<[^>]+>", " ", s or ""))).strip()


def city_of(address):
    if "大阪" not in (address or ""):
        return "大阪府内"
    m = re.search(r"大阪府\s*(大阪市[^\s0-9]*区|[^\s0-9]+市|[^\s0-9]+区|[^\s0-9]+町|[^\s0-9]+郡[^\s0-9]*)", address)
    return m.group(1) if m else "大阪府内"


def categorize(text):
    for kw in DELIVERY_KW:
        if kw in text:
            return "配送系"
    for kw in FACTORY_KW:
        if kw in text:
            return "工場系"
    for kw in WHITE_KW:
        if kw in text:
            return "ホワイトカラー"
    return "その他"


def parse_company_page(page_html):
    """/companies/{id} から 所在地・公式URL・事業内容・代表者・従業員数 を抽出。"""
    txt = clean(page_html)
    info = {"address": "", "url": "", "industry": "", "ceo": "", "employees": ""}
    m = re.search(r"大阪府[一-鿿0-9゠-ヿ\-－丁目番地号Ｆ0-9A-Za-z]{2,40}", txt)
    if m:
        info["address"] = m.group(0)
    # 公式サイトURL: まず「URL / ホームページ」ラベル直後を優先、無ければ本文から推定。
    EXCLUDE = (r"from-40\.jp|schema\.org|w3\.org|purl\.org|ogp\.me|gmpg\.org|"
               r"googleapis|gstatic|google\.com|googletagmanager|doubleclick|"
               r"facebook|twitter|x\.com|instagram|youtube|linkedin|line\.me|"
               r"cdn|jsdelivr|cloudflare|\.(png|jpg|jpeg|gif|svg|webp|css|js|ico)")

    def ok(u):
        return not re.search(EXCLUDE, u, re.I)

    lab = re.search(r'(?:URL|ホームページ|会社HP|コーポレートサイト)[^h]{0,20}(https?://[^\s"<)]+)', txt)
    if lab and ok(lab.group(1)):
        info["url"] = html.unescape(lab.group(1)).split("?")[0]
    else:
        for u in re.findall(r'href="(https?://[^"]+)"', page_html):
            if ok(u):
                info["url"] = html.unescape(u).split("?")[0]
                break
    mi = re.search(r"事業内容\s*[:：]?\s*(.+?)(?:URL|ホームページ|設立|資本金|代表者|従業員|この企業)", txt)
    if mi:
        info["industry"] = mi.group(1).strip()[:60]
    mc = re.search(r"代表者名?\s*[:：]?\s*([^\s設資従]{2,20})", txt)
    if mc:
        info["ceo"] = mc.group(1).strip()
    me = re.search(r"従業員数\s*[:：]?\s*([0-9,]+\s*人)", txt)
    if me:
        info["employees"] = me.group(1).strip()
    return info


def senior_flag(text):
    return "○" if any(k in text for k in SENIOR_KW) else "不明"


def collect(session, max_pages, max_companies, delay):
    companies = {}   # company_id -> {name, job, page_senior}
    order = []
    for page in range(1, max_pages + 1):
        url = f"{BASE}{LIST_PATH}&page={page}"
        try:
            page_html = fetch(session, url)
        except Exception as e:  # noqa
            print(f"  ! 一覧 page {page}: {e}", file=sys.stderr)
            break
        if not page_html:
            break
        cards = CARD_RE.findall(page_html)
        if not cards:
            print(f"  一覧 page {page}: 0件、終了")
            break
        page_senior = any(k in page_html for k in SENIOR_KW)
        added = 0
        for cid, cname, rid, jtitle in cards:
            cid = cid.strip()
            if cid in companies:
                continue
            companies[cid] = {
                "name": clean(cname), "job": clean(jtitle), "senior": page_senior,
            }
            order.append(cid)
            added += 1
        print(f"  一覧 page {page}: +{added}社 (計{len(companies)})")
        time.sleep(delay)
        if len(companies) >= max_companies:
            break

    # 会社ページで住所・公式URL・事業内容を補完
    rows = []
    targets = order[:max_companies]
    for i, cid in enumerate(targets, 1):
        c = companies[cid]
        info = {"address": "", "url": "", "industry": "", "ceo": "", "employees": ""}
        try:
            cp = fetch(session, f"{BASE}/companies/{cid}")
            if cp:
                info = parse_company_page(cp)
        except Exception as e:  # noqa
            print(f"  ! 会社 {cid}: {e}", file=sys.stderr)
        cat = categorize(f"{c['job']} {info['industry']}")
        memo_bits = [b for b in [info["industry"], info["address"],
                                 (f"代表 {info['ceo']}" if info["ceo"] else ""),
                                 (f"従業員 {info['employees']}" if info["employees"] else "")] if b]
        rows.append({
            "category": cat, "company": c["name"], "job": c["job"],
            "city": city_of(info["address"]),
            "senior": senior_flag(c["job"] + " " + ("シニア" if c["senior"] else "")),
            "url": info["url"], "memo": " / ".join(memo_bits),
        })
        if i % 20 == 0:
            print(f"  会社ページ補完 {i}/{len(targets)}")
        time.sleep(delay)
    return rows


def build_workbook(rows):
    by_cat = {c: [] for c in CATEGORIES}
    for r in rows:
        by_cat.get(r["category"], by_cat["その他"]).append(r)

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("リスト")
    for i, (h, w) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = w
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    idx = {h: i + 1 for i, (h, _) in enumerate(COLUMNS)}
    r, no = 2, 1
    for cat in CATEGORIES:
        for row in by_cat[cat]:
            vals = {
                "No.": no, "業種カテゴリ": cat, "企業名": row["company"],
                "募集職種": row["job"], "都道府県": "大阪府", "市区町村": row["city"],
                "掲載媒体": "FROM40", "シニア歓迎": row["senior"],
                "問い合わせURL/メール": row["url"], "ステータス": "未着手", "備考": row["memo"],
            }
            for h, _ in COLUMNS:
                cell = ws.cell(row=r, column=idx[h], value=vals.get(h, ""))
                cell.font = Font(name=FONT, size=10)
                cell.border = BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if h in ("No.", "業種カテゴリ", "都道府県"):
                    cell.fill = LABEL_FILL
            no += 1
            r += 1
    last = r - 1
    for h, opts in DV.items():
        dv = DataValidation(type="list", formula1=f'"{opts}"', allow_blank=True)
        ws.add_data_validation(dv)
        col = get_column_letter(idx[h])
        if last >= 2:
            dv.add(f"{col}2:{col}{last}")

    s = wb.create_sheet("説明", 0)
    s.sheet_view.showGridLines = False
    s.column_dimensions["A"].width = 20
    s.column_dimensions["B"].width = 84
    t = s.cell(row=1, column=1, value="RA営業リスト(大阪) — FROM40収集版")
    t.font = Font(name=FONT, bold=True, size=15, color="1F4E78")
    withurl = sum(1 for r in rows if r["url"])
    lines = [
        ("情報源", "FROM40(from-40.jp／40代・50代の転職) 大阪府×正社員。企業名・会社ページ公開。"),
        ("エリア/雇用", "大阪府 / 正社員（市区町村は会社所在地から抽出）。"),
        ("収集件数", f"{len(rows)}社（企業名で名寄せ・1社1行）"),
        ("連絡先", f"会社ページに電話番号の掲載は無いため、公式サイトURLを連絡導線として記載（{withurl}社取得）。架電先はURL先の公式サイトから取得してください。"),
        ("カテゴリ", "募集職種/事業内容から判定（配送→工場→ホワイトカラーの優先順、外れれば その他）。"),
        ("シニア歓迎", "40代・50代の転職媒体のため基本シニア層向け。求人タグに応じて○表記。"),
        ("留意", "掲載情報は変動します。架電前に各社公式・最新掲載で企業名/連絡先を再確認してください。"),
        ("参考", "senior-job.co.jp・sr.okjob.jp は求人一覧に雇用主企業名が出ないため収集対象外。"),
    ]
    rr = 3
    for k, v in lines:
        s.cell(row=rr, column=1, value=k).font = Font(name=FONT, bold=True, size=10)
        cc = s.cell(row=rr, column=2, value=v)
        cc.font = Font(name=FONT, size=10)
        cc.alignment = Alignment(wrap_text=True, vertical="top")
        rr += 1
    wb.active = 0
    return wb


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--max-pages", type=int, default=25)
    ap.add_argument("--max-companies", type=int, default=200)
    ap.add_argument("--delay", type=float, default=1.2)
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    session = requests.Session()
    session.headers.update({"User-Agent": UA, "Accept-Language": "ja,en;q=0.8"})
    ca = "/root/.ccr/ca-bundle.crt"
    if os.path.exists(ca):
        session.verify = ca

    print(f"== FROM40 大阪×正社員 収集 (delay {args.delay}s) ==")
    rows = collect(session, args.max_pages, args.max_companies, args.delay)

    out = args.out or os.path.join(os.path.dirname(__file__), "out",
                                   "RA営業リスト_大阪_FROM40.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb = build_workbook(rows)
    wb.save(out)
    from collections import Counter
    cc = Counter(r["category"] for r in rows)
    print(f"\nsaved: {out}  (計{len(rows)}社)")
    for c in CATEGORIES:
        print(f"  {c}: {cc.get(c, 0)}")


if __name__ == "__main__":
    main()
