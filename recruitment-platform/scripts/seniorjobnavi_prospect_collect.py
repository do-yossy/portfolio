#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
シニア求人ナビ(seniorjob-navi.com)の公開検索ページから、大阪府のシニア歓迎
求人を収集して RA(求人開拓)営業リスト(xlsx)に流し込む。

- 情報源: /zenkoku/PC27/EC1?page=N  (PC27=大阪府 / EC1=正社員)
  サーバーレンダリングHTML。robots.txt は未設置(全ページ取得可)だが、礼儀として
  自主待機を入れる。
- 各求人カードから 勤務先名・募集職種・勤務地・雇用形態・給与・電話番号・特徴タグ
  を抽出。電話は「直通番号」と「応募専用の転送番号(0078…)」を区別し、直通のみ
  電話欄へ、転送番号は備考へ。
- 企業名で名寄せし1社1行。職種からカテゴリを判定。

出力: scripts/out/RA営業リスト_大阪_シニア求人ナビ.xlsx

Usage:
  python3 scripts/seniorjobnavi_prospect_collect.py \
      [--white 100] [--delivery 50] [--factory 50] [--other 50] \
      [--max-pages 40] [--delay 2.0] [--out <path>]
"""
import argparse
import os
import re
import sys
import time
import html

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

BASE = "https://www.seniorjob-navi.com"
LIST_PATH = "/zenkoku/PC27/EC1"  # 大阪府 × 正社員
UA = "SeniorJobNavi-RA-list/1.0 (internal recruitment tool)"

DELIVERY_KW = ["ドライバー", "配送", "運搬", "配達", "トラック", "軽貨物", "ルート配",
               "宅配", "送迎", "運転手", "乗務員", "タクシー", "運転", "バス",
               "運行管理", "交通", "旅客", "配車", "運輸"]
FACTORY_KW = ["製造", "検品", "組立", "組み立て", "加工", "ライン", "梱包",
              "フォークリフト", "倉庫", "工場", "溶接", "機械オペ", "品質管理",
              "塗装", "プレス", "旋盤", "工員", "ものづくり", "包装", "整備"]
WHITE_KW = ["事務", "経理", "総務", "営業", "人事", "受付", "データ入力", "会計",
            "税理", "経営企画", "企画", "コールセンター", "テレア", "オペレーター",
            "CAD", "設計", "施工管理", "貿易", "秘書", "広報", "マーケ", "エンジニア",
            "プログラ", "行政書士", "社労士", "宅建", "士業", "コンサル",
            "管理業務", "総合職"]

SENIOR_TAGS = ("シニア", "50代", "60代", "70代", "年齢不問", "中高年", "ミドル", "年齢問わず")
CATEGORIES = ["ホワイトカラー", "配送系", "工場系", "その他"]

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
LABEL_FILL = PatternFill("solid", fgColor="D9E1F2")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = [
    ("No.", 6), ("業種カテゴリ", 14), ("企業名", 28), ("募集職種", 30),
    ("都道府県", 10), ("市区町村", 16), ("掲載媒体", 14), ("シニア歓迎", 10),
    ("電話番号", 16), ("問い合わせURL/メール", 24), ("掲載件数", 9),
    ("優先度", 8), ("ステータス", 12), ("送客候補メモ", 30), ("最終接触日", 12),
    ("備考", 26),
]
DV = {
    "業種カテゴリ": "ホワイトカラー,配送系,工場系,その他",
    "掲載媒体": "シニアジョブ,シニア求人ナビ,求人ボックス,Indeed,その他",
    "シニア歓迎": "○,△,×,不明",
    "優先度": "A,B,C",
    "ステータス": "未着手,架電済,アポ,送客,見送り,失注",
}


def fetch(session, page):
    url = f"{BASE}{LIST_PATH}"
    if page > 1:
        url += f"?page={page}"
    r = session.get(url, timeout=30)
    if r.status_code in (404, 410):
        return None
    r.raise_for_status()
    return r.text


def clean(s):
    return re.sub(r"\s+", " ", html.unescape(re.sub(r"<[^>]+>", " ", s))).strip()


def norm_phone(card_html, card_text):
    """(表示用電話, 種別) を返す。直通=('06-1234-5678','direct')、転送=('','tracking:...')"""
    # 表示テキストにハイフン付き番号があれば直通とみなす
    m = re.search(r"電話番号\d*\s*[:：]?\s*(0\d{1,3}-\d{1,4}-\d{3,4})", card_text)
    if m:
        return m.group(1), "direct"
    href = re.search(r'href="tel:([0-9\-]+)"', card_html)
    if not href:
        return "", ""
    d = re.sub(r"\D", "", href.group(1))
    # 応募専用の転送番号(0078600xxx…)や桁数異常はトラッキング扱い
    if d.startswith(("0078", "0570")) or len(d) not in (10, 11):
        return "", f"応募専用TEL {href.group(1)}"
    # 直通らしき10/11桁
    if len(d) == 10:
        disp = f"{d[:2]}-{d[2:6]}-{d[6:]}" if d[1] == "6" or d[1] == "3" else f"{d[:3]}-{d[3:6]}-{d[6:]}"
    else:
        disp = f"{d[:3]}-{d[3:7]}-{d[7:]}"
    return disp, "direct"


def city_of(location):
    if "大阪" not in (location or ""):
        return None
    m = re.search(r"大阪府\s*(大阪市[^\s0-9/]*区|[^\s0-9/]+市|[^\s0-9/]+区|[^\s0-9/]+町|[^\s0-9/]+郡[^\s0-9/]*)", location)
    return m.group(1) if m else "大阪府内"


def _match(text):
    for kw in DELIVERY_KW:
        if kw in text:
            return "配送系"
    for kw in FACTORY_KW:
        if kw in text:
            return "工場系"
    for kw in WHITE_KW:
        if kw in text:
            return "ホワイトカラー"
    return None


def categorize(title, fulltext):
    # 職種(タイトル)を優先。判定できなければ本文でフォールバック。
    return _match(title) or _match(fulltext) or "その他"


def senior_flag(text, tags):
    joined = text + " " + " ".join(tags)
    if any(t in joined for t in SENIOR_TAGS):
        return "○"
    if "ブランクOK" in joined:
        return "△"
    return "不明"


def parse_cards(page_html):
    blocks = re.split(r'(?=class="mod-jobResultBox plan001")', page_html)
    blocks = [b for b in blocks if "job-excerpt" in b and "勤務先名" in b]
    out = []
    for b in blocks:
        txt = clean(b)

        def grab(label, nxt):
            m = re.search(re.escape(label) + r"\s*(.+?)\s*(?:" + nxt + r")", txt)
            return m.group(1).strip() if m else ""

        company = grab("勤務先名", "勤務地")
        location = grab("勤務地", "雇用形態")
        emp = grab("雇用形態", "給与")
        sal = grab("給与", "応募資格|詳細をみる|電話応募")
        te = re.search(r'class="job-excerpt-wrap[^"]*"[^>]*>(.*?)</div>', b, re.S)
        title = clean(te.group(1)) if te else ""
        # 特徴タグ(勤務先名より前=ヘッダ部のアイコンラベル群)
        head = txt.split("勤務先名")[0]
        tags = [t for t in re.split(r"\s+", head) if t and t not in ("NEW",)]
        phone, ptype = norm_phone(b, txt)
        out.append({
            "company": company, "title": title, "location": location,
            "emp": emp, "sal": sal, "phone": phone, "ptype": ptype,
            "tags": tags, "fulltext": txt,
        })
    return out


def collect(session, targets, max_pages, delay):
    buckets = {c: [] for c in CATEGORIES}
    seen = set()
    for page in range(1, max_pages + 1):
        if all(len(buckets[c]) >= targets[c] for c in CATEGORIES):
            break
        try:
            page_html = fetch(session, page)
        except Exception as e:  # noqa
            print(f"  ! page {page}: {e}", file=sys.stderr)
            if "429" in str(e):
                time.sleep(delay * 4)
                continue
            break
        if page_html is None:
            break
        cards = parse_cards(page_html)
        if not cards:
            print(f"  page {page}: 0件、終了")
            break
        added = 0
        for c in cards:
            comp = c["company"]
            city = city_of(c["location"])
            if not comp or city is None:
                continue
            key = re.sub(r"\s+", "", comp)
            if key in seen:
                continue
            cat = categorize(c["title"], c["fulltext"])
            if len(buckets[cat]) >= targets[cat]:
                continue
            seen.add(key)
            memo = [x for x in [c["sal"], c["emp"]] if x]
            if c["ptype"] and c["ptype"] != "direct":
                memo.append(c["ptype"])
            if c["tags"]:
                memo.append("／".join(c["tags"][:5]))
            short = c["title"][:60] + ("…" if len(c["title"]) > 60 else "")
            buckets[cat].append({
                "company": comp, "title": short, "city": city,
                "phone": c["phone"], "senior": senior_flag(c["fulltext"], c["tags"]),
                "memo": " / ".join(memo),
            })
            added += 1
        counts = " ".join(f"{c[:2]}{len(buckets[c])}" for c in CATEGORIES)
        print(f"  page {page}: +{added} ({counts})")
        time.sleep(delay)
    return buckets


def build_workbook(buckets, targets):
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("リスト")
    for i, (h, w) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = w
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    idx = {h: i + 1 for i, (h, _) in enumerate(COLUMNS)}
    r, no = 2, 1
    for cat in CATEGORIES:
        for row in buckets.get(cat, []):
            vals = {
                "No.": no, "業種カテゴリ": cat, "企業名": row["company"],
                "募集職種": row["title"], "都道府県": "大阪府", "市区町村": row["city"],
                "掲載媒体": "シニア求人ナビ", "シニア歓迎": row["senior"],
                "電話番号": row["phone"], "ステータス": "未着手", "備考": row["memo"],
            }
            for h, _ in COLUMNS:
                cell = ws.cell(row=r, column=idx[h], value=vals.get(h, ""))
                cell.font = Font(name=FONT, size=10)
                cell.border = BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if h in ("No.", "業種カテゴリ", "都道府県"):
                    cell.fill = LABEL_FILL
            no += 1
            r += 1
    last = r - 1
    for h, opts in DV.items():
        dv = DataValidation(type="list", formula1=f'"{opts}"', allow_blank=True)
        ws.add_data_validation(dv)
        col = get_column_letter(idx[h])
        if last >= 2:
            dv.add(f"{col}2:{col}{last}")

    s = wb.create_sheet("説明", 0)
    s.sheet_view.showGridLines = False
    s.column_dimensions["A"].width = 20
    s.column_dimensions["B"].width = 84
    t = s.cell(row=1, column=1, value="RA営業リスト(大阪) — シニア求人ナビ収集版")
    t.font = Font(name=FONT, bold=True, size=15, color="1F4E78")
    tel_cnt = sum(1 for c in CATEGORIES for r in buckets.get(c, []) if r["phone"])
    total = sum(len(v) for v in buckets.values())
    lines = [
        ("情報源", "シニア求人ナビ(seniorjob-navi.com) /zenkoku/PC27/EC1(大阪府×正社員)。企業名公開の求人。"),
        ("エリア/雇用", "大阪府 / 正社員(市区町村は勤務地表記から抽出)。"),
        ("収集件数", " / ".join(f"{c}:{len(buckets.get(c, []))}/{targets[c]}" for c in CATEGORIES)),
        ("電話番号", f"直通番号を取得できたのは {tel_cnt}/{total} 社。応募専用の転送番号(0078…)は備考に記載。"),
        ("dedup", "企業名で名寄せし1社1行。"),
        ("カテゴリ", "募集職種/タグから判定(配送→工場→ホワイトカラーの優先順、外れれば その他)。"),
        ("シニア歓迎", "求人タグ/本文から判定(○=シニア/50〜70代/年齢不問、△=ブランクOK)。"),
        ("留意", "掲載情報は変動します。架電前に最新掲載で企業名/連絡先を再確認してください。"),
        ("参考", "リクナビNEXTはrobots.txtで当該検索(?prf=,?emp=,?l1=)が全面Disallowのため収集対象外。"),
    ]
    rr = 3
    for k, v in lines:
        s.cell(row=rr, column=1, value=k).font = Font(name=FONT, bold=True, size=10)
        cc = s.cell(row=rr, column=2, value=v)
        cc.font = Font(name=FONT, size=10)
        cc.alignment = Alignment(wrap_text=True, vertical="top")
        rr += 1
    wb.active = 0
    return wb


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--white", type=int, default=100)
    ap.add_argument("--delivery", type=int, default=50)
    ap.add_argument("--factory", type=int, default=50)
    ap.add_argument("--other", type=int, default=50)
    ap.add_argument("--max-pages", type=int, default=40)
    ap.add_argument("--delay", type=float, default=2.0)
    ap.add_argument("--out", default=None)
    args = ap.parse_args()
    targets = {"ホワイトカラー": args.white, "配送系": args.delivery,
               "工場系": args.factory, "その他": args.other}
    session = requests.Session()
    session.headers.update({"User-Agent": UA, "Accept-Language": "ja,en;q=0.8"})
    ca = "/root/.ccr/ca-bundle.crt"
    if os.path.exists(ca):
        session.verify = ca
    print(f"== シニア求人ナビ 大阪×正社員 収集 (delay {args.delay}s) ==")
    buckets = collect(session, targets, args.max_pages, args.delay)
    out = args.out or os.path.join(os.path.dirname(__file__), "out",
                                   "RA営業リスト_大阪_シニア求人ナビ.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb = build_workbook(buckets, targets)
    wb.save(out)
    total = sum(len(v) for v in buckets.values())
    print(f"\nsaved: {out}  (計{total}社)")
    for c in CATEGORIES:
        print(f"  {c}: {len(buckets[c])}")


if __name__ == "__main__":
    main()
