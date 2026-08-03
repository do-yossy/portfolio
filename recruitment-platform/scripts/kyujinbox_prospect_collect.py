#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
求人ボックス(kyujinbox)の公開検索ページから、大阪府の採用中企業を収集して
RA(求人開拓)営業リスト(xlsx)に流し込む。

- 情報源: 求人ボックス公開検索ページ /{keyword}の仕事-大阪府?pg=N
  (robots.txt で許可された検索/求人ページのみ。/rd/ 等の広告リダイレクトは辿らない)
- 企業名が表示される求人のみ対象。1社1行に集約(dedup)。
- カテゴリ別目標: ホワイトカラー100 / 配送系50 / 工場系50 (既定)。
- 出力: scripts/out/RA営業リスト_大阪_求人ボックス.xlsx

礼儀として自主的な待機(既定2秒)を入れ、ページ数に上限を設ける。
実行環境から求人ボックスへ到達できること(プロキシ/ネットワーク)が前提。

Usage:
  python3 scripts/kyujinbox_prospect_collect.py \
      [--white 100] [--delivery 50] [--factory 50] \
      [--max-pages 12] [--delay 2.0] [--out <path>]
"""
import argparse
import os
import re
import sys
import time
import html
from urllib.parse import quote

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

BASE = "https://xn--pckua2a7gp15o89zb.com"  # 求人ボックス.com (punycode)
UA = ("Mozilla/5.0 (compatible; RA-prospect-list/1.0; "
      "recruitment internal tool)")

# カテゴリ -> 検索キーワード(大阪府と組み合わせる)
CATEGORY_KEYWORDS = {
    "ホワイトカラー": ["事務", "経理", "総務", "営業事務", "一般事務",
                 "データ入力", "受付", "人事", "コールセンター"],
    "配送系": ["配送ドライバー", "ルート配送", "軽貨物", "宅配",
             "仕分け", "倉庫内作業", "ピッキング"],
    "工場系": ["製造", "検品", "組立", "フォークリフト", "梱包",
             "加工", "ライン作業"],
}

SENIOR_TAGS = ("シニア", "60代", "65歳", "年齢不問", "中高年", "ミドル", "定年")
BLANK_OK_TAGS = ("ブランクOK",)

# --- xlsx スタイル(テンプレートと共通) ---
FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
DATA_FILL = PatternFill("solid", fgColor="FFFFFF")
LABEL_FILL = PatternFill("solid", fgColor="D9E1F2")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = [
    ("No.", 6), ("業種カテゴリ", 14), ("企業名", 28), ("募集職種", 26),
    ("都道府県", 10), ("市区町村", 16), ("掲載媒体", 12), ("シニア歓迎", 10),
    ("電話番号", 16), ("問い合わせURL/メール", 26), ("掲載件数", 9),
    ("優先度", 8), ("ステータス", 12), ("送客候補メモ", 30), ("最終接触日", 12),
    ("備考", 24),
]
DV = {
    "業種カテゴリ": "ホワイトカラー,配送系,工場系",
    "掲載媒体": "シニアジョブ,求人ボックス,Indeed,スタンバイ,Engage,その他",
    "シニア歓迎": "○,△,×,不明",
    "優先度": "A,B,C",
    "ステータス": "未着手,架電済,アポ,送客,見送り,失注",
}


def fetch(session, path, pg):
    url = f"{BASE}{path}"
    if pg > 1:
        url += f"?pg={pg}"
    r = session.get(url, timeout=30)
    if r.status_code == 404:
        return None
    r.raise_for_status()
    return r.text


def clean(s):
    return re.sub(r"\s+", " ", html.unescape(re.sub(r"<[^>]+>", " ", s))).strip()


def parse_cards(page_html):
    """求人ボックス検索結果HTMLから求人カードを抽出。"""
    cards = re.split(r'(?=class="p-result_title--ver2")', page_html)
    out = []
    for c in cards:
        if "p-result_title_link" not in c:
            continue
        mt = re.search(r'class="p-result_title_link[^"]*"[^>]*>(.*?)</a>', c, re.S)
        mc = re.search(r'class="p-result_companyName"[^>]*>(.*?)</', c, re.S)
        title = clean(mt.group(1)) if mt else ""
        title = re.sub(r"\s*新着\s*$", "", title)
        company = clean(mc.group(1)) if mc else ""
        if not company:
            continue  # 企業名が出ない求人(広告等)は対象外
        # 勤務地 / 給与 / 雇用形態
        location = pay = emptype = ""
        cond = re.search(r'class="p-detail_company_conditions".*?</ul>', c, re.S)
        if cond:
            items = [clean(x) for x in re.findall(r"<li[^>]*>(.*?)</li>", cond.group(0), re.S)]
            items = [x for x in items if x]
            if items:
                location = items[0]
                if len(items) > 1:
                    pay = items[1]
                if len(items) > 2:
                    emptype = items[2]
        tags = [clean(t) for t in re.findall(
            r'class="p-result_tag_feature--ver2"[^>]*>(.*?)</', c, re.S)]
        tags = [t for t in tags if t]
        out.append({
            "company": company, "title": title, "location": location,
            "pay": pay, "emptype": emptype, "tags": tags,
        })
    return out


def city_of(location):
    """'大阪府 大阪市 大阪駅 / ...' -> '大阪市'。大阪以外は None。"""
    if "大阪" not in location:
        return None
    m = re.search(r"大阪府\s*([^\s/]+区|[^\s/]+市|[^\s/]+町|[^\s/]+郡[^\s/]*)", location)
    if m:
        return m.group(1)
    m = re.search(r"(大阪市[^\s/]*区|[^\s/]+市)", location)
    return m.group(1) if m else "大阪府内"


def senior_flag(tags):
    joined = " ".join(tags)
    if any(t in joined for t in SENIOR_TAGS):
        return "○"
    if any(t in joined for t in BLANK_OK_TAGS):
        return "△"
    return "不明"


def collect(session, category, target, max_pages, delay, seen):
    rows = []
    for kw in CATEGORY_KEYWORDS[category]:
        if len(rows) >= target:
            break
        path = "/" + quote(f"{kw}の仕事-大阪府", safe="")
        for pg in range(1, max_pages + 1):
            if len(rows) >= target:
                break
            try:
                page = fetch(session, path, pg)
            except Exception as e:  # noqa
                print(f"  ! {kw} pg{pg}: {e}", file=sys.stderr)
                break
            if page is None:
                break
            cards = parse_cards(page)
            if not cards:
                break
            added = 0
            for c in cards:
                if len(rows) >= target:
                    break
                city = city_of(c["location"])
                if city is None:
                    continue
                key = re.sub(r"\s+", "", c["company"])
                if key in seen:
                    continue
                seen.add(key)
                memo_bits = [b for b in [c["pay"], c["emptype"]] if b]
                if c["tags"]:
                    memo_bits.append("／".join(c["tags"][:6]))
                rows.append({
                    "category": category, "company": c["company"],
                    "title": c["title"], "city": city,
                    "senior": senior_flag(c["tags"]),
                    "memo": " / ".join(memo_bits),
                })
                added += 1
            print(f"  {category} '{kw}' pg{pg}: +{added} (計{len(rows)}/{target})")
            time.sleep(delay)
            if added == 0 and pg >= 2:
                break  # このキーワードは新規が枯れた
    return rows


def build_workbook(all_rows, targets):
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("リスト")
    for i, (h, w) in enumerate(COLUMNS, 1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = w
        cell = ws.cell(row=1, column=i)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    idx = {h: i + 1 for i, (h, _) in enumerate(COLUMNS)}

    r = 2
    no = 1
    for cat in ("ホワイトカラー", "配送系", "工場系"):
        rows = all_rows.get(cat, [])
        for row in rows:
            vals = {
                "No.": no, "業種カテゴリ": cat, "企業名": row["company"],
                "募集職種": row["title"], "都道府県": "大阪府", "市区町村": row["city"],
                "掲載媒体": "求人ボックス", "シニア歓迎": row["senior"],
                "ステータス": "未着手", "送客候補メモ": "", "備考": row["memo"],
            }
            for h, _ in COLUMNS:
                cell = ws.cell(row=r, column=idx[h], value=vals.get(h, ""))
                cell.font = Font(name=FONT, size=10)
                cell.border = BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if h in ("No.", "業種カテゴリ", "都道府県"):
                    cell.fill = LABEL_FILL
            no += 1
            r += 1
    last = r - 1

    for h, opts in DV.items():
        dv = DataValidation(type="list", formula1=f'"{opts}"', allow_blank=True)
        ws.add_data_validation(dv)
        col = get_column_letter(idx[h])
        if last >= 2:
            dv.add(f"{col}2:{col}{last}")

    # サマリタブ
    s = wb.create_sheet("説明", 0)
    s.sheet_view.showGridLines = False
    s.column_dimensions["A"].width = 20
    s.column_dimensions["B"].width = 80
    t = s.cell(row=1, column=1, value="RA営業リスト(大阪) — 求人ボックス収集版")
    t.font = Font(name=FONT, bold=True, size=15, color="1F4E78")
    lines = [
        ("情報源", "求人ボックス(kyujinbox) 公開検索ページ。企業名が表示される求人のみ収集。"),
        ("エリア", "大阪府(市区町村は勤務地表記から抽出)。"),
        ("収集件数", " / ".join(
            f"{c}:{len(all_rows.get(c, []))}/{targets[c]}" for c in
            ("ホワイトカラー", "配送系", "工場系"))),
        ("dedup", "企業名で名寄せし1社1行。"),
        ("シニア歓迎", "求人の特徴タグから推定(○=シニア/年齢不問系、△=ブランクOK、不明=記載なし)。要確認。"),
        ("留意", "掲載情報は変動します。架電前に各社公式・最新掲載で企業名/連絡先を再確認してください。"),
        ("連絡先", "電話・問い合わせURLは未取得(空欄)。各社公式サイト等から転記してください。"),
    ]
    rr = 3
    for k, v in lines:
        s.cell(row=rr, column=1, value=k).font = Font(name=FONT, bold=True, size=10)
        c = s.cell(row=rr, column=2, value=v)
        c.font = Font(name=FONT, size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        rr += 1
    wb.active = 0
    return wb


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--white", type=int, default=100)
    ap.add_argument("--delivery", type=int, default=50)
    ap.add_argument("--factory", type=int, default=50)
    ap.add_argument("--max-pages", type=int, default=12)
    ap.add_argument("--delay", type=float, default=2.0)
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    targets = {"ホワイトカラー": args.white, "配送系": args.delivery, "工場系": args.factory}

    session = requests.Session()
    session.headers.update({"User-Agent": UA, "Accept-Language": "ja,en;q=0.8"})
    ca = "/root/.ccr/ca-bundle.crt"
    if os.path.exists(ca):
        session.verify = ca

    seen = set()
    all_rows = {}
    for cat in ("ホワイトカラー", "配送系", "工場系"):
        print(f"== {cat} (目標{targets[cat]}) ==")
        all_rows[cat] = collect(session, cat, targets[cat], args.max_pages, args.delay, seen)

    out = args.out or os.path.join(os.path.dirname(__file__), "out",
                                   "RA営業リスト_大阪_求人ボックス.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb = build_workbook(all_rows, targets)
    wb.save(out)
    total = sum(len(v) for v in all_rows.values())
    print(f"\nsaved: {out}  (計{total}社)")


if __name__ == "__main__":
    main()
