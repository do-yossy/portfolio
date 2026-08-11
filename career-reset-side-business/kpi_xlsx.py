# -*- coding: utf-8 -*-
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = sys.argv[1] if len(sys.argv) > 1 else "kpi.xlsx"
WINE = "7A2233"; WINE_DK="591826"; BEIGE="F6EFE3"; GRAY="F3F1EF"; INK="23201F"; SUB="7C716B"
thin = Side(style="thin", color="D8CCC0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HFILL = PatternFill("solid", fgColor=WINE_DK)
HFONT = Font(name="游ゴシック", size=10.5, bold=True, color="FFFFFF")
CELL = Font(name="游ゴシック", size=10.5, color=INK)
TITLE = Font(name="游ゴシック", size=14, bold=True, color=WINE_DK)
NOTE = Font(name="游ゴシック", size=9, italic=True, color=SUB)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

wb = Workbook()

def header_row(ws, row, headers, widths):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = HFILL; c.font = HFONT; c.alignment = CENTER; c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = widths[i-1]
    ws.row_dimensions[row].height = 30

def style_body(ws, r0, r1, ncol):
    for r in range(r0, r1+1):
        for c in range(1, ncol+1):
            cell = ws.cell(row=r, column=c)
            cell.font = CELL; cell.border = BORDER; cell.alignment = LEFT
            if (r - r0) % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=BEIGE)

# ---------- Sheet 1: 見込み客管理 ----------
ws1 = wb.active; ws1.title = "見込み客管理"
ws1.cell(row=1, column=1, value="見込み客管理シート").font = TITLE
ws1.cell(row=2, column=1, value="Threads→LINE→無料相談→プログラムの流れを、1人ずつ記録します。ステータスと流入元はプルダウンから選べます。").font = NOTE
heads1 = ["No.","名前／ニックネーム","流入元","LINE追加日","特典送付","初回フォロー日","相談予約日","相談実施日","ステータス","成約日","金額（円）","メモ"]
w1 = [5,18,10,12,9,13,12,12,12,12,12,26]
header_row(ws1, 4, heads1, w1)
for i in range(5, 45):
    ws1.cell(row=i, column=1, value=i-4)
style_body(ws1, 5, 44, len(heads1))
ws1.freeze_panes = "A5"
# dropdowns
dv_status = DataValidation(type="list", formula1='"新規,特典送付済,フォロー中,相談予約,相談実施,成約,見送り"', allow_blank=True)
dv_src = DataValidation(type="list", formula1='"Threads,Instagram,紹介,その他"', allow_blank=True)
dv_yn = DataValidation(type="list", formula1='"済,未"', allow_blank=True)
ws1.add_data_validation(dv_status); ws1.add_data_validation(dv_src); ws1.add_data_validation(dv_yn)
dv_status.add("I5:I44"); dv_src.add("C5:C44"); dv_yn.add("E5:E44")
for r in range(5,45):
    ws1.cell(row=r, column=11).number_format = '#,##0'

# ---------- Sheet 2: 週次KPI ----------
ws2 = wb.create_sheet("週次KPI")
ws2.cell(row=1, column=1, value="週次KPIボード").font = TITLE
ws2.cell(row=2, column=1, value="毎週の数字を入れると、右側の『通過率』が自動計算されます。ボトルネック（追加数・相談化率）を毎週改善しましょう。").font = NOTE
heads2 = ["週","投稿数","インプレッション","プロフクリック","LINE追加","相談予約","相談実施","成約","売上（円）","プロフクリック率","追加率","相談化率","成約率"]
w2 = [12,8,14,12,10,10,10,8,12,12,10,10,10]
header_row(ws2, 4, heads2, w2)
weeks = ["1週目","2週目","3週目","4週目","5週目","6週目","7週目","8週目"]
r0 = 5
for k, wk in enumerate(weeks):
    r = r0 + k
    ws2.cell(row=r, column=1, value=wk)
    # formulas: プロフクリック率=クリック/インプ, 追加率=追加/クリック, 相談化率=相談実施/追加, 成約率=成約/相談実施
    ws2.cell(row=r, column=10, value=f'=IFERROR(D{r}/C{r},"")')
    ws2.cell(row=r, column=11, value=f'=IFERROR(E{r}/D{r},"")')
    ws2.cell(row=r, column=12, value=f'=IFERROR(G{r}/E{r},"")')
    ws2.cell(row=r, column=13, value=f'=IFERROR(H{r}/G{r},"")')
rN = r0 + len(weeks)
ws2.cell(row=rN, column=1, value="合計/平均")
for col in [2,3,4,5,6,7,8,9]:
    L = get_column_letter(col)
    ws2.cell(row=rN, column=col, value=f'=SUM({L}{r0}:{L}{rN-1})')
ws2.cell(row=rN, column=10, value=f'=IFERROR(D{rN}/C{rN},"")')
ws2.cell(row=rN, column=11, value=f'=IFERROR(E{rN}/D{rN},"")')
ws2.cell(row=rN, column=12, value=f'=IFERROR(G{rN}/E{rN},"")')
ws2.cell(row=rN, column=13, value=f'=IFERROR(H{rN}/G{rN},"")')
style_body(ws2, r0, rN, len(heads2))
for r in range(r0, rN+1):
    ws2.cell(row=r, column=9).number_format = '#,##0'
    for c in [10,11,12,13]:
        ws2.cell(row=r, column=c).number_format = '0.0%'
# emphasize total row
for c in range(1, len(heads2)+1):
    cell = ws2.cell(row=rN, column=c); cell.font = Font(name="游ゴシック", size=10.5, bold=True, color=INK); cell.fill = PatternFill("solid", fgColor=GRAY)
ws2.freeze_panes = "A5"

# ---------- Sheet 3: 目標逆算 ----------
ws3 = wb.create_sheet("目標逆算")
ws3.cell(row=1, column=1, value="目標からの逆算（必要な数）").font = TITLE
ws3.cell(row=2, column=1, value="黄色いセル（目標成約数・各通過率）を変えると、必要な相談・LINE追加・プロフクリック数が自動計算されます。").font = NOTE
IN = PatternFill("solid", fgColor="FFF3C4")
def kv(row, label, value, fmt=None, is_input=False, formula=None):
    lc = ws3.cell(row=row, column=1, value=label); lc.font = Font(name="游ゴシック", size=10.5, bold=True, color=INK); lc.border=BORDER; lc.alignment=LEFT
    vc = ws3.cell(row=row, column=2, value=(formula if formula else value)); vc.font = CELL; vc.border=BORDER; vc.alignment=CENTER
    if fmt: vc.number_format = fmt
    if is_input: vc.fill = IN
    return vc
ws3.column_dimensions['A'].width = 34; ws3.column_dimensions['B'].width = 16
kv(4, "月の目標 成約数（件）", 3, '0', True)
kv(5, "相談→成約の通過率", 0.30, '0%', True)
kv(6, "LINE追加→相談 の通過率", 0.30, '0%', True)
kv(7, "プロフクリック→LINE追加 の通過率", 0.25, '0%', True)
kv(8, "営業日数（月）", 20, '0', True)
kv(10, "必要な 無料相談（件/月）", None, '0.0', formula='=IFERROR(B4/B5,"")')
kv(11, "必要な LINE追加（件/月）", None, '0.0', formula='=IFERROR(B10/B6,"")')
kv(12, "必要な プロフクリック（件/月）", None, '0.0', formula='=IFERROR(B11/B7,"")')
kv(13, "1営業日あたり プロフクリック", None, '0.0', formula='=IFERROR(B12/B8,"")')
ws3.cell(row=15, column=1, value="※通過率は仮の目安です。実測（週次KPI）に合わせて数字を書き換えてください。成果を保証するものではありません。").font = NOTE

wb.save(OUT)
print("WROTE", OUT)
